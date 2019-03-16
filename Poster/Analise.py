import openpyxl
from openpyxl.styles import PatternFill
import datetime
import numpy as np

class AnaliseClients:
    def __init__(self, clients) -> None:
        self.__clients = clients
        self.RFM = RFM(self.__clients)


class RFM:
    def __init__(self, clients) -> None:
        self.__clients = clients
        self.dates = [client.last_order_date_timestamp for client in self.__clients]
        self.count_orders = [client.count_order for client in self.__clients]
        self.payed_sum = [client.total_payed_sum for client in self.__clients]
        self.clients_rfm = self.make()

    def make(self):
        clients_rfm = []
        for client in self.__clients:
            client.recency = self.recency(client)
            client.frequency = self.frequency(client)
            client.monetary = self.monetary(client)
            client.rfm = f"{client.recency}{client.frequency}{client.monetary}"
            clients_rfm.append(client)

        return clients_rfm

    def recency(self, client):
        date_max = datetime.datetime.fromtimestamp(np.percentile(self.dates, 66) / 1e3).date()
        date_min = datetime.datetime.fromtimestamp(np.percentile(self.dates, 33) / 1e3).date()

        if client.last_order_date > date_max:
            return 3
        elif client.last_order_date > date_min:
            return 2
        return 1

    def frequency(self, client):
        count_max = np.percentile(self.count_orders, 87)
        count_min = np.percentile(self.count_orders, 70)

        if client.count_order > count_max:
            return 3
        elif client.count_order > count_min:
            return 2
        return 1

    def monetary(self, client):
        payed_max = np.percentile(self.payed_sum, 50)
        payed_min = np.percentile(self.payed_sum, 25)

        if client.total_payed_sum > payed_max:
            return 3
        elif client.total_payed_sum > payed_min:
            return 2
        return 1

    @staticmethod
    def write_in_sheet(sheet, clients):
        sheet['A1'] = 'Клиент'
        sheet['B1'] = 'Номер телефона'
        sheet['C1'] = 'Максимум по полю закритий'
        sheet['D1'] = 'Количество по полю '
        sheet['E1'] = 'Сумма по полю'
        sheet['F1'] = 'R'
        sheet['G1'] = 'F'
        sheet['H1'] = 'M'
        sheet['I1'] = 'RFM'

        sheet.column_dimensions['A'].width = 20
        sheet.column_dimensions['B'].width = 20
        sheet.column_dimensions['C'].width = 25
        sheet.column_dimensions['D'].width = 17
        sheet.column_dimensions['E'].width = 13

        sheet['A1'].fill = PatternFill(start_color="009900", end_color="009900", fill_type="solid")
        sheet['B1'].fill = PatternFill(start_color="00cc00", end_color="00cc00", fill_type="solid")
        sheet['C1'].fill = PatternFill(start_color="ff0000", end_color="ff0000", fill_type="solid")
        sheet['D1'].fill = PatternFill(start_color="fdff00", end_color="fdff00", fill_type="solid")
        sheet['E1'].fill = PatternFill(start_color="03a9f3", end_color="03a9f3", fill_type="solid")
        sheet['F1'].fill = PatternFill(start_color="ff0000", end_color="ff0000", fill_type="solid")
        sheet['G1'].fill = PatternFill(start_color="fdff00", end_color="fdff00", fill_type="solid")
        sheet['H1'].fill = PatternFill(start_color="03a9f3", end_color="03a9f3", fill_type="solid")
        sheet['I1'].fill = PatternFill(start_color="d0ffa0", end_color="d0ffa0", fill_type="solid")

        row = 2
        for client in clients:
            sheet.cell(row=row, column=1).value = client.name
            sheet.cell(row=row, column=2).value = client.phone_number
            sheet.cell(row=row, column=3).value = client.last_order_date
            sheet.cell(row=row, column=4).value = client.count_order
            sheet.cell(row=row, column=5).value = client.total_payed_sum
            sheet.cell(row=row, column=6).value = client.recency
            sheet.cell(row=row, column=7).value = client.frequency
            sheet.cell(row=row, column=8).value = client.monetary
            sheet.cell(row=row, column=9).value = client.rfm

            sheet.cell(row=row, column=3).number_format = 'mm.dd.yyyy'
            row += 1

    def export_to_excel(self, path_to_save_and_file_name):
        doc = openpyxl.Workbook()
        ALL = doc.get_active_sheet()
        ALL.title = 'ALL'

        self.write_in_sheet(ALL, self.clients_rfm)
        for sheet in sorted(set([client.rfm for client in self.clients_rfm])):
            self.write_in_sheet(doc.create_sheet(sheet), list(filter(lambda x: x.rfm == sheet, self.clients_rfm)))

        doc.save(path_to_save_and_file_name)

        return path_to_save_and_file_name
